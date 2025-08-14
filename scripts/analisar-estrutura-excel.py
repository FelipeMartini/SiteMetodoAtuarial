#!/usr/bin/env python3
"""
Script para analisar a estrutura do arquivo Excel de mortalidade
Extrai informações sobre planilhas, colunas, dados e estrutura
"""

import sys
import os
import openpyxl
from openpyxl import load_workbook
import json
from datetime import datetime

def analisar_estrutura_excel(caminho_arquivo):
    """
    Analisa a estrutura completa de um arquivo Excel
    Retorna informações sobre planilhas, colunas, dados e estrutura
    """
    
    print(f"🔍 ANÁLISE ESTRUTURAL DO ARQUIVO EXCEL")
    print(f"📁 Arquivo: {os.path.basename(caminho_arquivo)}")
    print(f"📊 Tamanho: {os.path.getsize(caminho_arquivo):,} bytes")
    print("=" * 80)
    
    try:
        # Carrega o arquivo Excel
        workbook = load_workbook(caminho_arquivo, data_only=False)
        
        resultado = {
            "arquivo": os.path.basename(caminho_arquivo),
            "tamanho_bytes": os.path.getsize(caminho_arquivo),
            "data_analise": datetime.now().isoformat(),
            "planilhas": []
        }
        
        print(f"📋 PLANILHAS ENCONTRADAS: {len(workbook.sheetnames)}")
        
        for idx, nome_planilha in enumerate(workbook.sheetnames, 1):
            print(f"\n{idx}. PLANILHA: '{nome_planilha}'")
            print("-" * 50)
            
            worksheet = workbook[nome_planilha]
            
            # Dimensões da planilha
            max_row = worksheet.max_row
            max_col = worksheet.max_column
            
            print(f"   📐 Dimensões: {max_row} linhas × {max_col} colunas")
            print(f"   📍 Intervalo: A1:{worksheet.cell(max_row, max_col).coordinate}")
            
            # Analisa cabeçalhos (primeira linha)
            print(f"\n   📝 CABEÇALHOS (Linha 1):")
            cabecalhos = []
            for col in range(1, min(max_col + 1, 21)):  # Máximo 20 colunas para evitar spam
                celula = worksheet.cell(1, col)
                valor = celula.value
                cabecalhos.append({
                    "coluna": celula.coordinate.replace('1', ''),
                    "valor": str(valor) if valor is not None else "",
                    "tipo": type(valor).__name__
                })
                print(f"      {celula.coordinate}: {valor} ({type(valor).__name__})")
            
            # Analisa algumas linhas de dados
            print(f"\n   🔢 AMOSTRA DE DADOS (Linhas 2-6):")
            linhas_amostra = []
            for row in range(2, min(max_row + 1, 7)):  # Linhas 2-6
                linha_dados = []
                print(f"      Linha {row}:")
                for col in range(1, min(max_col + 1, 11)):  # Máximo 10 colunas
                    celula = worksheet.cell(row, col)
                    valor = celula.value
                    linha_dados.append({
                        "celula": celula.coordinate,
                        "valor": valor,
                        "tipo": type(valor).__name__
                    })
                    print(f"         {celula.coordinate}: {valor} ({type(valor).__name__})")
                linhas_amostra.append(linha_dados)
            
            # Busca por dados específicos de mortalidade
            print(f"\n   🎯 BUSCA POR DADOS DE MORTALIDADE:")
            termos_busca = ['qx', 'mort', 'idade', 'óbito', 'participante', 'massa', 'tábua', 'exp', 'obs']
            celulas_relevantes = []
            
            for row in range(1, min(max_row + 1, 101)):  # Primeiras 100 linhas
                for col in range(1, min(max_col + 1, 21)):  # Primeiras 20 colunas
                    celula = worksheet.cell(row, col)
                    valor_str = str(celula.value).lower() if celula.value else ""
                    
                    if any(termo in valor_str for termo in termos_busca):
                        celulas_relevantes.append({
                            "celula": celula.coordinate,
                            "valor": celula.value,
                            "linha": row,
                            "coluna": col
                        })
                        print(f"      ✓ {celula.coordinate}: {celula.value}")
            
            # Adiciona informações da planilha ao resultado
            resultado["planilhas"].append({
                "nome": nome_planilha,
                "dimensoes": {
                    "linhas": max_row,
                    "colunas": max_col,
                    "intervalo": f"A1:{worksheet.cell(max_row, max_col).coordinate}"
                },
                "cabecalhos": cabecalhos,
                "linhas_amostra": linhas_amostra,
                "celulas_relevantes": celulas_relevantes
            })
        
        print(f"\n" + "=" * 80)
        print(f"✅ ANÁLISE CONCLUÍDA")
        print(f"📊 Total de planilhas: {len(resultado['planilhas'])}")
        
        return resultado
        
    except Exception as e:
        print(f"❌ ERRO na análise: {str(e)}")
        return None

def main():
    # Caminho do arquivo Excel
    caminho_arquivo = "/home/felipe/Área de Trabalho/GitHub/SiteMetodoAtuarial/revisao-completa/MORTALIDADE APOSENTADOS dez 2024 2019 A 2024 FELIPE qx masc e fem (Massa Janeiro).xlsx"
    
    if not os.path.exists(caminho_arquivo):
        print(f"❌ Arquivo não encontrado: {caminho_arquivo}")
        return 1
    
    # Executa a análise
    resultado = analisar_estrutura_excel(caminho_arquivo)
    
    if resultado:
        # Salva o resultado em JSON
        arquivo_resultado = "/home/felipe/Área de Trabalho/GitHub/SiteMetodoAtuarial/XLOGS/analise-estrutura-excel.json"
        with open(arquivo_resultado, 'w', encoding='utf-8') as f:
            json.dump(resultado, f, indent=2, ensure_ascii=False, default=str)
        
        print(f"💾 Resultado salvo em: {arquivo_resultado}")
        return 0
    else:
        return 1

if __name__ == "__main__":
    sys.exit(main())
