#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script Python para análise avançada de arquivos Excel de mortalidade
Compatível com OpenPyXL, Pandas, NumPy e SciPy
"""

import sys
import json
import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from scipy import stats
import warnings
from datetime import datetime, date
import re
from typing import Dict, List, Any, Optional, Union

# Configurar warnings
warnings.filterwarnings('ignore', category=UserWarning)

class AnalisadorMortalidadeExcel:
    def __init__(self, caminho_arquivo: str, configuracao: Dict[str, Any]):
        self.caminho_arquivo = caminho_arquivo
        self.configuracao = configuracao
        self.workbook = None
        self.dados_extraidos = {}
        self.estatisticas = {}
        self.validacao = {'erros_encontrados': [], 'warnings': [], 'integridade_ok': True}
        self.formulas_encontradas = []
        
    def carregar_arquivo(self):
        """Carrega o arquivo Excel usando OpenPyXL e Pandas"""
        try:
            # Carregar com OpenPyXL para preservar fórmulas
            self.workbook = load_workbook(self.caminho_arquivo, data_only=False)
            
            # Listar todas as planilhas
            planilhas = self.workbook.sheetnames
            
            # Filtrar planilhas específicas se configurado
            if self.configuracao.get('planilhasEspecificas'):
                planilhas_filtradas = []
                for nome in self.configuracao['planilhasEspecificas']:
                    planilhas_encontradas = [p for p in planilhas if nome.lower() in p.lower()]
                    planilhas_filtradas.extend(planilhas_encontradas)
                planilhas = planilhas_filtradas if planilhas_filtradas else planilhas
                
            return planilhas
            
        except Exception as e:
            self.validacao['erros_encontrados'].append(f"Erro ao carregar arquivo: {str(e)}")
            raise
    
    def extrair_formulas(self, worksheet):
        """Extrai fórmulas da planilha"""
        formulas = []
        
        if not self.configuracao.get('extrairFormulas', True):
            return formulas
            
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.data_type == 'f' and cell.value:  # Fórmula
                    formulas.append({
                        'celula': cell.coordinate,
                        'formula': cell.value,
                        'valor_calculado': cell.value if worksheet.parent.calculation.value else None
                    })
        
        return formulas
    
    def analisar_planilha_massa(self, worksheet) -> Dict[str, Any]:
        """Analisa planilha de massa de participantes"""
        try:
            # Converter para DataFrame
            data = []
            headers = None
            
            for i, row in enumerate(worksheet.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(cell).strip() if cell else f"col_{j}" for j, cell in enumerate(row)]
                    continue
                    
                if any(cell is not None for cell in row):
                    data.append(row)
            
            if not data:
                return {'erro': 'Planilha vazia'}
                
            df = pd.DataFrame(data, columns=headers)
            
            # Identificar colunas importantes
            colunas_identificadas = self.identificar_colunas(df.columns)
            
            # Estatísticas básicas
            estatisticas = {
                'total_registros': len(df),
                'colunas_identificadas': colunas_identificadas,
                'registros_validos': 0,
                'registros_com_erro': 0
            }
            
            # Analisar dados por coluna identificada
            dados_processados = {}
            
            for tipo_coluna, nome_coluna in colunas_identificadas.items():
                if nome_coluna and nome_coluna in df.columns:
                    serie = df[nome_coluna].dropna()
                    
                    if tipo_coluna == 'idade':
                        dados_processados['distribuicao_idade'] = self.analisar_distribuicao_idade(serie)
                    elif tipo_coluna == 'sexo':
                        dados_processados['distribuicao_sexo'] = self.analisar_distribuicao_sexo(serie)
                    elif tipo_coluna == 'data_nascimento':
                        dados_processados['analise_datas'] = self.analisar_datas(serie)
                    elif tipo_coluna == 'salario':
                        dados_processados['estatisticas_salario'] = self.analisar_valores_monetarios(serie)
            
            # Contar registros válidos
            estatisticas['registros_validos'] = len(df.dropna(subset=[col for col in colunas_identificadas.values() if col]))
            estatisticas['registros_com_erro'] = estatisticas['total_registros'] - estatisticas['registros_validos']
            
            return {
                'estatisticas': estatisticas,
                'dados_processados': dados_processados,
                'amostra_dados': df.head(5).to_dict('records')
            }
            
        except Exception as e:
            return {'erro': f'Erro ao analisar massa: {str(e)}'}
    
    def analisar_planilha_obitos(self, worksheet) -> Dict[str, Any]:
        """Analisa planilha de óbitos"""
        try:
            # Converter para DataFrame
            data = []
            headers = None
            
            for i, row in enumerate(worksheet.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(cell).strip() if cell else f"col_{j}" for j, cell in enumerate(row)]
                    continue
                    
                if any(cell is not None for cell in row):
                    data.append(row)
            
            if not data:
                return {'erro': 'Planilha de óbitos vazia'}
                
            df = pd.DataFrame(data, columns=headers)
            
            # Identificar colunas
            colunas_identificadas = self.identificar_colunas_obitos(df.columns)
            
            # Análise específica de óbitos
            estatisticas_obitos = {
                'total_obitos': len(df),
                'colunas_identificadas': colunas_identificadas,
                'distribuicao_por_idade': {},
                'distribuicao_por_sexo': {},
                'distribuicao_temporal': {}
            }
            
            # Analisar distribuição por idade nos óbitos
            if colunas_identificadas.get('idade_obito'):
                col_idade = colunas_identificadas['idade_obito']
                if col_idade in df.columns:
                    idades = df[col_idade].dropna()
                    estatisticas_obitos['distribuicao_por_idade'] = self.analisar_distribuicao_idade(idades)
            
            # Analisar distribuição por sexo nos óbitos
            if colunas_identificadas.get('sexo'):
                col_sexo = colunas_identificadas['sexo']
                if col_sexo in df.columns:
                    sexos = df[col_sexo].dropna()
                    estatisticas_obitos['distribuicao_por_sexo'] = self.analisar_distribuicao_sexo(sexos)
            
            # Analisar distribuição temporal
            if colunas_identificadas.get('data_obito'):
                col_data = colunas_identificadas['data_obito']
                if col_data in df.columns:
                    datas = df[col_data].dropna()
                    estatisticas_obitos['distribuicao_temporal'] = self.analisar_distribuicao_temporal(datas)
            
            return {
                'estatisticas': estatisticas_obitos,
                'amostra_dados': df.head(5).to_dict('records')
            }
            
        except Exception as e:
            return {'erro': f'Erro ao analisar óbitos: {str(e)}'}
    
    def analisar_planilha_qx(self, worksheet) -> Dict[str, Any]:
        """Analisa planilha de qx (taxas de mortalidade)"""
        try:
            # Converter para DataFrame
            data = []
            headers = None
            
            for i, row in enumerate(worksheet.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(cell).strip() if cell else f"col_{j}" for j, cell in enumerate(row)]
                    continue
                    
                if any(cell is not None for cell in row):
                    data.append(row)
            
            if not data:
                return {'erro': 'Planilha qx vazia'}
                
            df = pd.DataFrame(data, columns=headers)
            
            # Identificar colunas de qx
            colunas_qx = self.identificar_colunas_qx(df.columns)
            
            # Análise das taxas qx
            estatisticas_qx = {
                'total_idades': len(df),
                'colunas_identificadas': colunas_qx,
                'faixa_idades': {},
                'estatisticas_qx': {}
            }
            
            # Analisar faixa de idades
            if colunas_qx.get('idade'):
                col_idade = colunas_qx['idade']
                if col_idade in df.columns:
                    idades = df[col_idade].dropna()
                    estatisticas_qx['faixa_idades'] = {
                        'idade_minima': int(idades.min()),
                        'idade_maxima': int(idades.max()),
                        'total_idades': len(idades)
                    }
            
            # Analisar taxas qx por sexo
            for sexo in ['masculino', 'feminino']:
                col_qx = colunas_qx.get(f'qx_{sexo}')
                if col_qx and col_qx in df.columns:
                    qx_values = pd.to_numeric(df[col_qx], errors='coerce').dropna()
                    if len(qx_values) > 0:
                        estatisticas_qx['estatisticas_qx'][sexo] = {
                            'media': float(qx_values.mean()),
                            'mediana': float(qx_values.median()),
                            'desvio_padrao': float(qx_values.std()),
                            'minimo': float(qx_values.min()),
                            'maximo': float(qx_values.max()),
                            'total_valores': len(qx_values)
                        }
            
            return {
                'estatisticas': estatisticas_qx,
                'amostra_dados': df.head(5).to_dict('records')
            }
            
        except Exception as e:
            return {'erro': f'Erro ao analisar qx: {str(e)}'}
    
    def identificar_colunas(self, colunas: List[str]) -> Dict[str, str]:
        """Identifica automaticamente as colunas importantes"""
        mapeamento = {
            'idade': None,
            'sexo': None,
            'data_nascimento': None,
            'salario': None,
            'nome': None,
            'cpf': None
        }
        
        for col in colunas:
            col_lower = col.lower().strip()
            
            # Identificar idade
            if any(termo in col_lower for termo in ['idade', 'age', 'anos']):
                mapeamento['idade'] = col
            
            # Identificar sexo
            elif any(termo in col_lower for termo in ['sexo', 'sex', 'genero', 'gender']):
                mapeamento['sexo'] = col
            
            # Identificar data de nascimento
            elif any(termo in col_lower for termo in ['nascimento', 'birth', 'data_nasc', 'dt_nasc']):
                mapeamento['data_nascimento'] = col
            
            # Identificar salário
            elif any(termo in col_lower for termo in ['salario', 'salary', 'remuneracao', 'valor']):
                mapeamento['salario'] = col
            
            # Identificar nome
            elif any(termo in col_lower for termo in ['nome', 'name', 'participante']):
                mapeamento['nome'] = col
            
            # Identificar CPF
            elif any(termo in col_lower for termo in ['cpf', 'documento', 'id']):
                mapeamento['cpf'] = col
        
        return mapeamento
    
    def identificar_colunas_obitos(self, colunas: List[str]) -> Dict[str, str]:
        """Identifica colunas específicas de óbitos"""
        mapeamento = {
            'idade_obito': None,
            'sexo': None,
            'data_obito': None,
            'causa_obito': None,
            'cpf': None
        }
        
        for col in colunas:
            col_lower = col.lower().strip()
            
            if any(termo in col_lower for termo in ['idade', 'age']):
                mapeamento['idade_obito'] = col
            elif any(termo in col_lower for termo in ['sexo', 'sex']):
                mapeamento['sexo'] = col
            elif any(termo in col_lower for termo in ['data', 'obito', 'death', 'falecimento']):
                mapeamento['data_obito'] = col
            elif any(termo in col_lower for termo in ['causa', 'motivo', 'reason']):
                mapeamento['causa_obito'] = col
            elif any(termo in col_lower for termo in ['cpf', 'documento', 'id']):
                mapeamento['cpf'] = col
        
        return mapeamento
    
    def identificar_colunas_qx(self, colunas: List[str]) -> Dict[str, str]:
        """Identifica colunas de taxas qx"""
        mapeamento = {
            'idade': None,
            'qx_masculino': None,
            'qx_feminino': None,
            'qx_geral': None
        }
        
        for col in colunas:
            col_lower = col.lower().strip()
            
            if any(termo in col_lower for termo in ['idade', 'age', 'x']):
                mapeamento['idade'] = col
            elif any(termo in col_lower for termo in ['qx', 'q(x)', 'taxa']) and any(termo in col_lower for termo in ['masc', 'male', 'm']):
                mapeamento['qx_masculino'] = col
            elif any(termo in col_lower for termo in ['qx', 'q(x)', 'taxa']) and any(termo in col_lower for termo in ['fem', 'female', 'f']):
                mapeamento['qx_feminino'] = col
            elif any(termo in col_lower for termo in ['qx', 'q(x)', 'taxa']) and not any(termo in col_lower for termo in ['masc', 'male', 'm', 'fem', 'female', 'f']):
                mapeamento['qx_geral'] = col
        
        return mapeamento
    
    def analisar_distribuicao_idade(self, serie_idade) -> Dict[str, Any]:
        """Analisa distribuição de idades"""
        idades = pd.to_numeric(serie_idade, errors='coerce').dropna()
        
        if len(idades) == 0:
            return {'erro': 'Nenhuma idade válida encontrada'}
        
        return {
            'total': len(idades),
            'media': float(idades.mean()),
            'mediana': float(idades.median()),
            'desvio_padrao': float(idades.std()),
            'minima': int(idades.min()),
            'maxima': int(idades.max()),
            'distribuicao_faixas': {
                '0-20': len(idades[idades <= 20]),
                '21-30': len(idades[(idades > 20) & (idades <= 30)]),
                '31-40': len(idades[(idades > 30) & (idades <= 40)]),
                '41-50': len(idades[(idades > 40) & (idades <= 50)]),
                '51-60': len(idades[(idades > 50) & (idades <= 60)]),
                '61-70': len(idades[(idades > 60) & (idades <= 70)]),
                '71+': len(idades[idades > 70])
            }
        }
    
    def analisar_distribuicao_sexo(self, serie_sexo) -> Dict[str, Any]:
        """Analisa distribuição por sexo"""
        sexos = serie_sexo.dropna().astype(str).str.upper()
        
        # Normalizar valores
        sexos_normalizados = sexos.map({
            'M': 'MASCULINO', 'MASC': 'MASCULINO', 'MASCULINO': 'MASCULINO', 'MALE': 'MASCULINO',
            'F': 'FEMININO', 'FEM': 'FEMININO', 'FEMININO': 'FEMININO', 'FEMALE': 'FEMININO'
        }).fillna('OUTROS')
        
        distribuicao = sexos_normalizados.value_counts().to_dict()
        total = len(sexos_normalizados)
        
        return {
            'total': total,
            'distribuicao': distribuicao,
            'percentuais': {k: round(v/total*100, 2) for k, v in distribuicao.items()}
        }
    
    def analisar_datas(self, serie_datas) -> Dict[str, Any]:
        """Analisa série de datas"""
        try:
            datas = pd.to_datetime(serie_datas, errors='coerce').dropna()
            
            if len(datas) == 0:
                return {'erro': 'Nenhuma data válida encontrada'}
            
            return {
                'total': len(datas),
                'data_minima': datas.min().strftime('%Y-%m-%d'),
                'data_maxima': datas.max().strftime('%Y-%m-%d'),
                'distribuicao_por_ano': datas.dt.year.value_counts().to_dict(),
                'distribuicao_por_mes': datas.dt.month.value_counts().to_dict()
            }
        except Exception:
            return {'erro': 'Erro ao processar datas'}
    
    def analisar_valores_monetarios(self, serie_valores) -> Dict[str, Any]:
        """Analisa valores monetários"""
        valores = pd.to_numeric(serie_valores, errors='coerce').dropna()
        
        if len(valores) == 0:
            return {'erro': 'Nenhum valor monetário válido encontrado'}
        
        return {
            'total': len(valores),
            'media': float(valores.mean()),
            'mediana': float(valores.median()),
            'desvio_padrao': float(valores.std()),
            'minimo': float(valores.min()),
            'maximo': float(valores.max()),
            'quartis': {
                'q1': float(valores.quantile(0.25)),
                'q2': float(valores.quantile(0.5)),
                'q3': float(valores.quantile(0.75))
            }
        }
    
    def analisar_distribuicao_temporal(self, serie_datas) -> Dict[str, Any]:
        """Analisa distribuição temporal de óbitos"""
        try:
            datas = pd.to_datetime(serie_datas, errors='coerce').dropna()
            
            if len(datas) == 0:
                return {'erro': 'Nenhuma data de óbito válida encontrada'}
            
            return {
                'total_obitos': len(datas),
                'periodo': {
                    'inicio': datas.min().strftime('%Y-%m-%d'),
                    'fim': datas.max().strftime('%Y-%m-%d')
                },
                'distribuicao_anual': datas.dt.year.value_counts().sort_index().to_dict(),
                'distribuicao_mensal': datas.dt.month.value_counts().sort_index().to_dict(),
                'tendencia': self.calcular_tendencia_temporal(datas)
            }
        except Exception as e:
            return {'erro': f'Erro ao analisar distribuição temporal: {str(e)}'}
    
    def calcular_tendencia_temporal(self, datas) -> Dict[str, Any]:
        """Calcula tendência temporal dos óbitos"""
        try:
            # Agrupar por ano e contar
            obitos_por_ano = datas.dt.year.value_counts().sort_index()
            
            if len(obitos_por_ano) < 2:
                return {'erro': 'Dados insuficientes para calcular tendência'}
            
            anos = obitos_por_ano.index.values
            counts = obitos_por_ano.values
            
            # Regressão linear simples
            slope, intercept, r_value, p_value, std_err = stats.linregress(anos, counts)
            
            return {
                'coeficiente_angular': float(slope),
                'intercepto': float(intercept),
                'correlacao': float(r_value),
                'r_quadrado': float(r_value**2),
                'p_value': float(p_value),
                'erro_padrao': float(std_err),
                'interpretacao': 'crescente' if slope > 0 else 'decrescente' if slope < 0 else 'estável'
            }
        except Exception:
            return {'erro': 'Erro no cálculo de tendência'}
    
    def executar_analise(self) -> Dict[str, Any]:
        """Executa análise completa do arquivo"""
        resultado = {
            'metadados': {
                'arquivo': os.path.basename(self.caminho_arquivo),
                'tamanho_arquivo': os.path.getsize(self.caminho_arquivo),
                'processado_em': datetime.now().isoformat(),
                'configuracao_utilizada': self.configuracao,
                'versaoPython': sys.version,
                'bibliotecas': {
                    'pandas': pd.__version__,
                    'numpy': np.__version__,
                    'openpyxl': '3.1.2'  # versão estimada
                }
            },
            'estrutura_arquivo': {},
            'dados_extraidos': {},
            'estatisticas': {},
            'validacao': self.validacao,
            'warnings': []
        }
        
        try:
            # Carregar arquivo
            planilhas = self.carregar_arquivo()
            
            resultado['estrutura_arquivo'] = {
                'planilhas': planilhas,
                'total_planilhas': len(planilhas)
            }
            
            total_linhas = 0
            formulas_total = 0
            
            # Analisar cada planilha
            for nome_planilha in planilhas:
                worksheet = self.workbook[nome_planilha]
                
                # Contar linhas
                linhas_planilha = worksheet.max_row
                total_linhas += linhas_planilha
                
                # Extrair fórmulas se configurado
                if self.configuracao.get('extrairFormulas', True):
                    formulas = self.extrair_formulas(worksheet)
                    formulas_total += len(formulas)
                    self.formulas_encontradas.extend(formulas)
                
                # Identificar tipo de planilha e analisar
                nome_lower = nome_planilha.lower()
                
                if any(termo in nome_lower for termo in ['massa', 'participante', 'trabalhada', 'unificada']):
                    resultado['dados_extraidos']['massa_participantes'] = self.analisar_planilha_massa(worksheet)
                
                elif any(termo in nome_lower for termo in ['obito', 'morte', 'falecimento', 'death']):
                    resultado['dados_extraidos']['obitos_registrados'] = self.analisar_planilha_obitos(worksheet)
                
                elif any(termo in nome_lower for termo in ['qx', 'mortalidade', 'taxa']):
                    resultado['dados_extraidos']['qx_mortalidade'] = self.analisar_planilha_qx(worksheet)
            
            # Estatísticas gerais
            resultado['estrutura_arquivo']['total_linhas'] = total_linhas
            resultado['estrutura_arquivo']['formulas_encontradas'] = formulas_total
            
            # Calcular estatísticas consolidadas
            resultado['estatisticas'] = self.calcular_estatisticas_consolidadas()
            
            # Validação final
            if not resultado['dados_extraidos']:
                self.validacao['warnings'].append('Nenhum dado específico foi extraído das planilhas')
            
            resultado['validacao'] = self.validacao
            resultado['warnings'] = self.validacao['warnings']
            
            return resultado
            
        except Exception as e:
            self.validacao['erros_encontrados'].append(f"Erro na análise: {str(e)}")
            self.validacao['integridade_ok'] = False
            resultado['validacao'] = self.validacao
            return resultado
    
    def calcular_estatisticas_consolidadas(self) -> Dict[str, Any]:
        """Calcula estatísticas consolidadas de todos os dados"""
        stats = {
            'resumo_geral': {
                'massa_analisada': False,
                'obitos_analisados': False,
                'qx_analisado': False,
                'total_formulas': len(self.formulas_encontradas)
            }
        }
        
        # Verificar massa
        if 'massa_participantes' in self.dados_extraidos:
            massa = self.dados_extraidos['massa_participantes']
            if 'estatisticas' in massa:
                stats['resumo_geral']['massa_analisada'] = True
                stats['massa_estatisticas'] = massa['estatisticas']
        
        # Verificar óbitos
        if 'obitos_registrados' in self.dados_extraidos:
            obitos = self.dados_extraidos['obitos_registrados']
            if 'estatisticas' in obitos:
                stats['resumo_geral']['obitos_analisados'] = True
                stats['obitos_estatisticas'] = obitos['estatisticas']
        
        # Verificar qx
        if 'qx_mortalidade' in self.dados_extraidos:
            qx = self.dados_extraidos['qx_mortalidade']
            if 'estatisticas' in qx:
                stats['resumo_geral']['qx_analisado'] = True
                stats['qx_estatisticas'] = qx['estatisticas']
        
        return stats

def main():
    """Função principal do script"""
    if len(sys.argv) != 3:
        print(json.dumps({
            'erro': 'Uso: python script.py <caminho_arquivo> <configuracao_json>',
            'argumentos_recebidos': sys.argv
        }))
        sys.exit(1)
    
    try:
        caminho_arquivo = sys.argv[1]
        configuracao_json = sys.argv[2]
        
        # Parse da configuração
        configuracao = json.loads(configuracao_json)
        
        # Verificar se arquivo existe
        if not os.path.exists(caminho_arquivo):
            print(json.dumps({
                'erro': f'Arquivo não encontrado: {caminho_arquivo}'
            }))
            sys.exit(1)
        
        # Executar análise
        analisador = AnalisadorMortalidadeExcel(caminho_arquivo, configuracao)
        resultado = analisador.executar_analise()
        
        # Retornar resultado como JSON
        print(json.dumps(resultado, ensure_ascii=False, indent=2))
        
    except json.JSONDecodeError as e:
        print(json.dumps({
            'erro': f'Erro ao fazer parse da configuração JSON: {str(e)}'
        }))
        sys.exit(1)
    
    except Exception as e:
        print(json.dumps({
            'erro': f'Erro inesperado: {str(e)}',
            'tipo_erro': type(e).__name__
        }))
        sys.exit(1)

if __name__ == '__main__':
    main()
